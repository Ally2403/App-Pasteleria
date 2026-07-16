import openpyxl
import uuid

JHON_UUID = '18f2014a-92ab-4ddd-9a36-5ba83c7d1ea1'
PACKAGING_KW = ['caja', 'blonda', 'capacillo', 'bolsa', 'opalina', 'topper', 'bandeja', 'papel']
LABOR_KW = ['mano de obra']
SERVICE_KW = ['servicios', 'servicio']

def classify(name):
    if not name: return 'ingredient'
    n = str(name).lower().strip()
    for kw in LABOR_KW:
        if kw in n: return 'labor'
    for kw in SERVICE_KW:
        if kw in n: return 'service'
    for kw in PACKAGING_KW:
        if kw in n: return 'packaging'
    return 'ingredient'

def num(v, d=0):
    try: return float(v) if v is not None else d
    except: return d

def esc(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

ingredients_map = {}

def get_or_create_ingredient(name, lugar, qty_sold, unit, price_raw):
    key = str(name).lower().strip()
    if key not in ingredients_map:
        qty_sold_f = num(qty_sold, 1)
        price_f = num(price_raw, 0)
        unit_price = price_f / qty_sold_f if qty_sold_f > 0 else 0
        ingredients_map[key] = {
            'id': str(uuid.uuid4()),
            'name': str(name).strip(),
            'provider': str(lugar).strip() if lugar else None,
            'qty_sold': qty_sold_f,
            'unit': str(unit).strip() if unit else 'unidad',
            'price': price_f,
            'unit_price': unit_price
        }
    return ingredients_map[key]['id']

all_recipes = []

# ─── VITALI ──────────────────────────────────────────────────────────────────
vitali_file = r'c:\Users\Allison Ruiz\Desktop\app mami\EXCELS\COSTOS VITALI.xlsx'
wb_v = openpyxl.load_workbook(vitali_file, data_only=True)

for shname in wb_v.sheetnames:
    ws = wb_v[shname]
    rows = list(ws.iter_rows(values_only=True))

    recipe_name = None
    for row in rows:
        if row[1] and isinstance(row[1], str) and len(row[1]) > 5 and row[1] == row[1].upper():
            recipe_name = row[1].strip()
            break
    if not recipe_name:
        continue

    recipe_id = str(uuid.uuid4())
    ing_list, extra_list = [], []
    in_data = False
    total_cost = None
    rounded_price = None

    for row in rows:
        if row[1] == 'Lugar' and row[2] == 'Ingredientes':
            in_data = True
            continue
        if in_data:
            if row[1] and isinstance(row[1], str) and 'Total costo' in row[1]:
                total_cost = num(row[8])
                in_data = False
                continue
            name = row[2]
            if not name:
                continue
            lugar = row[1]
            qty_sold = row[3]
            unit = row[4]
            price_raw = row[5]
            qty_used = row[6]
            unit2 = row[7]
            item_type = classify(name)
            if item_type == 'ingredient':
                ing_id = get_or_create_ingredient(name, lugar, qty_sold, unit or unit2, price_raw)
                ing_list.append({'ingredient_id': ing_id, 'qty_used': num(qty_used)})
            else:
                qty_sold_f = num(qty_sold, 1)
                price_f = num(price_raw, 0)
                unit_cost = price_f / qty_sold_f if qty_sold_f > 0 else price_f
                extra_list.append({
                    'name': str(name).strip(),
                    'type': item_type,
                    'quantity': num(qty_used),
                    'unit_price': unit_cost,
                    'total': num(row[8])
                })

        if row[1] and isinstance(row[1], str) and '50%' in row[1]:
            rounded_price = num(row[9]) if row[9] else num(row[8])

    cost_per_unit = (total_cost / 20) if total_cost else 0
    suggested = cost_per_unit * 1.5
    if not rounded_price:
        rounded_price = round(suggested / 100) * 100

    all_recipes.append({
        'id': recipe_id,
        'name': recipe_name,
        'description': 'Vitali - receta con socio Jhon',
        'units_per_batch': 20,
        'has_partner': True,
        'partner_id': JHON_UUID,
        'ingredients': ing_list,
        'extra_costs': extra_list,
        'pricing': {
            'profit_percentage': 50.0,
            'suggested_price': round(suggested, 2),
            'rounded_price': rounded_price
        }
    })

# ─── ALLY ────────────────────────────────────────────────────────────────────
ally_file = r'c:\Users\Allison Ruiz\Desktop\app mami\EXCELS\COSTOS ALLY BAKERY.xlsx'
wb_a = openpyxl.load_workbook(ally_file, data_only=True)
ws_a = wb_a['Torta vainilla']
rows_a = list(ws_a.iter_rows(values_only=True))

BLOCKS = [
    {'title_row': 1,  'data_start': 5,  'data_end': 19, 'total_row': 20},
    {'title_row': 22, 'data_start': 26, 'data_end': 40, 'total_row': 41},
    {'title_row': 43, 'data_start': 47, 'data_end': 61, 'total_row': 62},
]

# Horizontal sections: (lugar_col, name_col, qty_sold_col, unit_col, price_col, qty_used_col, unit2_col, total_col)
SECTION_COLS = [
    (1, 2, 3, 4, 5, 6, 7, 8),
    (10, 11, 12, 13, 14, 15, 16, 17),
    (19, 20, 21, 22, 23, 24, 25, 26),
]

SKIP_NAMES = {'ingredientes', 'lugar', 'vendo en', 'ganancia', 'total', 'total con cosas adicionales',
              'total costo 2 pisos', 'total costo 3 pisos', 'total costo pequena', 'total torta 3 pisos',
              'total torta pequena'}

for block in BLOCKS:
    title_row = rows_a[block['title_row']]
    recipe_name_raw = str(title_row[1]).strip() if title_row[1] else 'TORTA VAINILLA'
    recipe_id = str(uuid.uuid4())
    ing_list, extra_list = [], []
    vendo_en = None

    for ri in range(block['data_start'], block['data_end'] + 1):
        row = rows_a[ri]
        for sc in SECTION_COLS:
            lugar_i, name_i, qty_sold_i, unit_i, price_i, qty_used_i, unit2_i, total_i = sc
            if total_i >= len(row):
                continue
            name = row[name_i]
            if not name or not isinstance(name, str):
                continue
            if name.lower().strip() in SKIP_NAMES:
                continue
            total_val = row[total_i]
            if not isinstance(total_val, (int, float)):
                continue

            lugar = row[lugar_i]
            qty_sold = row[qty_sold_i]
            unit = row[unit_i]
            price_raw = row[price_i]
            qty_used = row[qty_used_i]
            unit2 = row[unit2_i]

            item_type = classify(name)
            if item_type == 'ingredient':
                ing_id = get_or_create_ingredient(name, lugar, qty_sold, unit or unit2, price_raw)
                ing_list.append({'ingredient_id': ing_id, 'qty_used': num(qty_used)})
            else:
                qty_sold_f = num(qty_sold, 1)
                price_f = num(price_raw, 0)
                unit_cost = price_f / qty_sold_f if qty_sold_f > 0 else price_f
                extra_list.append({
                    'name': str(name).strip(),
                    'type': item_type,
                    'quantity': num(qty_used),
                    'unit_price': unit_cost,
                    'total': num(total_val)
                })

        # Check for VENDO EN in this row
        for col_idx in [26, 17, 8]:
            if col_idx >= len(row): continue
            cell = row[col_idx]
            if isinstance(cell, (int, float)) and cell > 10000:
                prev = row[col_idx - 1] if col_idx > 0 else None
                if prev and isinstance(prev, str) and 'vendo en' in str(prev).lower():
                    vendo_en = num(cell)

    total_row = rows_a[block['total_row']]
    total_cost = num(total_row[8])

    if vendo_en and total_cost > 0:
        profit_pct = round(((vendo_en - total_cost) / total_cost) * 100, 2)
        suggested = vendo_en
        rounded_p = vendo_en
    else:
        profit_pct = 50.0
        suggested = round(total_cost * 1.5, 2)
        rounded_p = round(suggested / 1000) * 1000

    all_recipes.append({
        'id': recipe_id,
        'name': recipe_name_raw,
        'description': 'Ally Bakery - Torta de Vainilla',
        'units_per_batch': 1,
        'has_partner': False,
        'partner_id': None,
        'ingredients': ing_list,
        'extra_costs': extra_list,
        'pricing': {
            'profit_percentage': profit_pct,
            'suggested_price': round(suggested, 2),
            'rounded_price': rounded_p
        }
    })

# ─── SUMMARY ─────────────────────────────────────────────────────────────────
print(f'Ingredientes unicos: {len(ingredients_map)}')
print(f'Recetas: {len(all_recipes)}')
for r in all_recipes:
    nm = r['name']
    up = r['units_per_batch']
    hp = r['has_partner']
    ni = len(r['ingredients'])
    ne = r['extra_costs']
    rp = r['pricing']['rounded_price']
    print(f'  [{nm}] | {up} uds | partner={hp} | {ni} ings | {len(ne)} extras | precio={rp}')

print('\nIngredientes unicos registrados:')
for k, v in ingredients_map.items():
    print(f"  {v['name']} | {v['unit']} | unit_price={v['unit_price']:.4f}")

# ─── GENERATE SQL ────────────────────────────────────────────────────────────
lines = []
lines.append('-- ====================================================')
lines.append('-- DATOS INICIALES - App Pasteleria Mami')
lines.append('-- Generado automaticamente desde los excels')
lines.append('-- ====================================================')
lines.append('')

# Fix Jhon role
lines.append('-- 1. CORREGIR ROL DE JHON (de admin a partner)')
lines.append(f"UPDATE public.profiles SET role = 'partner' WHERE id = '{JHON_UUID}';")
lines.append('')

# Ingredients
lines.append('-- 2. INGREDIENTES')
for v in ingredients_map.values():
    lines.append(
        f"INSERT INTO ingredients (id, name, provider, quantity_sold, unit, price, unit_price) VALUES "
        f"({esc(v['id'])}, {esc(v['name'])}, {esc(v['provider'])}, {v['qty_sold']}, {esc(v['unit'])}, {v['price']}, {round(v['unit_price'], 6)});"
    )
lines.append('')

# Inventory
lines.append('-- 3. INVENTARIO (stock inicial en 0)')
for v in ingredients_map.values():
    lines.append(
        f"INSERT INTO inventory (ingredient_id, current_stock, unit) VALUES "
        f"({esc(v['id'])}, 0, {esc(v['unit'])});"
    )
lines.append('')

# Recipes
lines.append('-- 4. RECETAS')
for r in all_recipes:
    hp = 'true' if r['has_partner'] else 'false'
    pid = esc(r['partner_id'])
    lines.append(
        f"INSERT INTO recipes (id, name, description, units_per_batch, has_partner, partner_id) VALUES "
        f"({esc(r['id'])}, {esc(r['name'])}, {esc(r['description'])}, {r['units_per_batch']}, {hp}, {pid});"
    )
lines.append('')

# Recipe ingredients
lines.append('-- 5. INGREDIENTES POR RECETA')
for r in all_recipes:
    for ing in r['ingredients']:
        lines.append(
            f"INSERT INTO recipe_ingredients (recipe_id, ingredient_id, quantity_used) VALUES "
            f"({esc(r['id'])}, {esc(ing['ingredient_id'])}, {ing['qty_used']});"
        )
lines.append('')

# Recipe extra costs
lines.append('-- 6. COSTOS EXTRA POR RECETA')
for r in all_recipes:
    for ex in r['extra_costs']:
        lines.append(
            f"INSERT INTO recipe_extra_costs (recipe_id, name, type, quantity, unit_price, total) VALUES "
            f"({esc(r['id'])}, {esc(ex['name'])}, {esc(ex['type'])}, {ex['quantity']}, {round(ex['unit_price'], 4)}, {round(ex['total'], 2)});"
        )
lines.append('')

# Recipe pricing
lines.append('-- 7. PRECIOS')
for r in all_recipes:
    p = r['pricing']
    lines.append(
        f"INSERT INTO recipe_pricing (recipe_id, profit_percentage, suggested_price, rounded_price) VALUES "
        f"({esc(r['id'])}, {p['profit_percentage']}, {p['suggested_price']}, {p['rounded_price']});"
    )

sql_output = '\n'.join(lines)
out_path = r'c:\Users\Allison Ruiz\Desktop\app mami\schemas\supabase_seed_data.sql'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(sql_output)

print(f'\nSQL generado: {out_path}')
print(f'Total lineas: {len(lines)}')
